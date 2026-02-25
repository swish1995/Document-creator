"""ExcelLoader 동적 시트명 테스트 (Phase 2.1, 2.2)"""

import json
import logging
import pytest
from pathlib import Path
from unittest.mock import patch, MagicMock

import openpyxl


@pytest.fixture
def excel_with_custom_sheet(tmp_path):
    """커스텀 시트명이 있는 엑셀 파일"""
    file_path = tmp_path / "custom.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MyData"
    ws.append(["Name", "Value"])
    ws.append(["Item1", 100])
    ws.append(["Item2", 200])
    wb.save(file_path)
    wb.close()
    return file_path


@pytest.fixture
def excel_with_capture_data(tmp_path):
    """"Capture Data" 시트가 있는 엑셀 파일"""
    file_path = tmp_path / "capture.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Capture Data"
    ws.append(["Col1", "Col2"])
    ws.append(["A", 1])
    wb.save(file_path)
    wb.close()
    return file_path


@pytest.fixture
def excel_with_multiple_sheets(tmp_path):
    """여러 시트가 있는 엑셀 파일"""
    file_path = tmp_path / "multi.xlsx"
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1.append(["H1", "H2"])
    ws1.append(["S1V1", "S1V2"])

    ws2 = wb.create_sheet("Sheet2")
    ws2.append(["X1", "X2"])
    ws2.append(["S2V1", "S2V2"])
    wb.save(file_path)
    wb.close()
    return file_path


class TestExcelLoaderDynamicSheet:
    """ExcelLoader 동적 시트명 테스트"""

    def test_load_with_explicit_sheet_name(self, excel_with_custom_sheet, tmp_path):
        """명시적 시트명으로 로드"""
        from src.core.excel_loader import ExcelLoader

        loader = ExcelLoader(base_dir=tmp_path)
        loader.load(excel_with_custom_sheet, sheet_name="MyData")

        headers = loader.get_headers()
        assert "Name" in headers
        assert "Value" in headers
        assert loader.row_count == 2

    def test_load_with_default_captures_data(self, excel_with_capture_data, tmp_path):
        """시트명 미지정 시 'Capture Data' 시트 사용 (기존 동작)"""
        from src.core.excel_loader import ExcelLoader

        loader = ExcelLoader(base_dir=tmp_path)
        loader.load(excel_with_capture_data)

        headers = loader.get_headers()
        assert "Col1" in headers

    def test_load_fallback_to_first_sheet(self, excel_with_custom_sheet, tmp_path):
        """'Capture Data' 시트 없으면 첫 번째 시트 사용"""
        from src.core.excel_loader import ExcelLoader

        loader = ExcelLoader(base_dir=tmp_path)
        loader.load(excel_with_custom_sheet)

        headers = loader.get_headers()
        assert "Name" in headers

    def test_load_nonexistent_sheet_raises_error(self, excel_with_custom_sheet, tmp_path):
        """존재하지 않는 시트명 → ExcelLoaderError"""
        from src.core.excel_loader import ExcelLoader, ExcelLoaderError

        loader = ExcelLoader(base_dir=tmp_path)
        with pytest.raises(ExcelLoaderError, match="시트를 찾을 수 없습니다"):
            loader.load(excel_with_custom_sheet, sheet_name="NonExistent")

    def test_load_specific_sheet_from_multiple(self, excel_with_multiple_sheets, tmp_path):
        """여러 시트 중 특정 시트 선택"""
        from src.core.excel_loader import ExcelLoader

        loader = ExcelLoader(base_dir=tmp_path)
        loader.load(excel_with_multiple_sheets, sheet_name="Sheet2")

        headers = loader.get_headers()
        assert "X1" in headers
        assert "X2" in headers


class TestFormulaEvaluatorOptional:
    """수식 평가기 선택적 활성화 테스트"""

    def test_load_plain_excel_without_named_ranges(self, excel_with_custom_sheet, tmp_path):
        """Named Range 없는 일반 엑셀을 에러 없이 로드"""
        from src.core.excel_loader import ExcelLoader

        loader = ExcelLoader(base_dir=tmp_path)
        loader.load(excel_with_custom_sheet)

        # 에러 없이 로드되고 데이터 접근 가능
        assert loader.is_loaded
        assert loader.row_count == 2
        row = loader.get_row(0)
        assert row["Name"] == "Item1"
        assert row["Value"] == 100

    def test_log_warning_when_skipping_formulas(self, excel_with_custom_sheet, tmp_path, caplog):
        """Named Range 없는 엑셀 로드 시 수식 평가 스킵 관련 로그"""
        from src.core.excel_loader import ExcelLoader

        loader = ExcelLoader(base_dir=tmp_path)
        with caplog.at_level(logging.INFO):
            loader.load(excel_with_custom_sheet)

        log_messages = " ".join(caplog.messages)
        assert "수식 평가 스킵" in log_messages
