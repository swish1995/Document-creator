"""엑셀 파일 로더 모듈

Skeleton Analyzer에서 생성된 엑셀 파일을 로드하고 데이터를 제공합니다.
수식이 포함된 셀은 내장 수식 평가기를 사용하여 계산합니다.
이미지가 포함된 셀은 임시 디렉토리에 추출하여 경로를 제공합니다.
"""

from __future__ import annotations

import re
import shutil
import time
import traceback
from concurrent.futures import ThreadPoolExecutor, as_completed
from io import BytesIO
from pathlib import Path
from typing import Any, Optional, List, Dict, Union, Tuple

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.exceptions import InvalidFileException
from PIL import Image

from src.core.logger import get_logger

# 셀 참조 패턴: 영문 1~3자 + 숫자 (앞에 영문/밑줄이 없어야 함)
_CELL_REF_RE = re.compile(r"(?<![A-Za-z_])[A-Z]{1,3}\d+")
# 등호(=) → 이중등호(==) 변환 (<=, >=, <> 제외)
_EQ_RE = re.compile(r"(?<![<>])=(?!=)")
_UNRESOLVED = "__UNRESOLVED__"


def _excel_if(condition, true_val, false_val):
    """Excel IF 함수"""
    return true_val if condition else false_val


def _excel_index(table, row, col=None):
    """Excel INDEX 함수 (1-based)"""
    row = int(row)
    if col is None:
        return table[row - 1][0]
    col = int(col)
    return table[row - 1][col - 1]


class ExcelLoaderError(Exception):
    """ExcelLoader 관련 에러"""

    pass


class ExcelLoader:
    """엑셀 파일 로더

    Skeleton Analyzer 출력 엑셀 파일의 Capture Data 시트를 로드합니다.
    """

    DEFAULT_SHEET_NAME = "Capture Data"
    IMAGES_DIR_NAME = ".images"
    THUMBNAIL_SIZE = 40  # 썸네일 크기

    def __init__(self, base_dir: Path | str | None = None):
        self._logger = get_logger("excel_loader")
        self._workbook = None
        self._sheet = None
        self._headers: list[str] = []
        self._data: list[dict[str, Any]] = []
        self._data_by_index: list[list[Any]] = []  # 인덱스 기반 데이터
        self._file_path: Path | None = None
        self._calculated_values: dict = {}  # 수식 계산 결과 캐시
        self._use_data_only: bool = False  # data_only 폴백 모드 여부
        self._image_map: dict[str, Path] = {}  # 셀 주소 → 이미지 경로 매핑
        self._thumbnail_map: dict[str, Path] = {}  # 셀 주소 → 썸네일 경로 매핑

        # 이미지 디렉토리 설정
        if base_dir is None:
            base_dir = Path(__file__).parent.parent.parent
        self._base_dir = Path(base_dir)
        self._images_dir = self._base_dir / self.IMAGES_DIR_NAME

    @property
    def images_dir(self) -> Path:
        """이미지 디렉토리 경로"""
        return self._images_dir

    @property
    def thumbnail_map(self) -> dict[str, Path]:
        """썸네일 경로 매핑"""
        return self._thumbnail_map.copy()

    def get_image_path(self, row: int, col: int) -> Path | None:
        """특정 셀의 이미지 경로 반환

        Args:
            row: 행 인덱스 (0부터 시작, 데이터 행 기준)
            col: 열 인덱스 (0부터 시작)

        Returns:
            이미지 경로, 없으면 None
        """
        # 엑셀 행은 1-based, 헤더가 1행이므로 데이터는 2행부터
        cell_key = f"{row + 2}_{col}"
        return self._image_map.get(cell_key)

    def get_thumbnail_path(self, row: int, col: int) -> Path | None:
        """특정 셀의 썸네일 경로 반환

        Args:
            row: 행 인덱스 (0부터 시작, 데이터 행 기준)
            col: 열 인덱스 (0부터 시작)

        Returns:
            썸네일 경로, 없으면 None
        """
        cell_key = f"{row + 2}_{col}"
        return self._thumbnail_map.get(cell_key)

    def cleanup_images(self) -> None:
        """이미지 디렉토리 정리"""
        if self._images_dir.exists():
            shutil.rmtree(self._images_dir)
            self._logger.info(f"이미지 디렉토리 삭제: {self._images_dir}")

    @property
    def is_loaded(self) -> bool:
        """데이터가 로드되었는지 여부"""
        return self._workbook is not None

    @property
    def file_path(self) -> Path | None:
        """로드된 파일 경로"""
        return self._file_path

    @property
    def row_count(self) -> int:
        """전체 행 수"""
        self._ensure_loaded()
        return len(self._data)

    def load(self, file_path: Path | str, progress_callback: callable = None) -> None:
        """엑셀 파일 로드

        Args:
            file_path: 엑셀 파일 경로
            progress_callback: 진행 상황 콜백 (step: int, message: str)
                - step 1: 엑셀 파일 읽기 중
                - step 2: 수식 계산 중
                - step 3: 값 변환 중
                - step 4: 이미지 추출 중
                - step 5: 데이터 로드 중

        Raises:
            ExcelLoaderError: 파일을 찾을 수 없거나 형식이 잘못된 경우
        """
        file_path = Path(file_path)

        if not file_path.exists():
            raise ExcelLoaderError(f"파일을 찾을 수 없습니다: {file_path}")

        # 기존 이미지 디렉토리 정리
        self.cleanup_images()

        def update_progress(step: int, message: str):
            if progress_callback:
                progress_callback(step, message)

        try:
            # 1~3단계: 내장 수식 평가기로 수식 계산
            self._logger.info(f"수식 계산 시작: {file_path}")
            self._calculate_formulas(file_path, update_progress)
            self._logger.info("수식 계산 완료")

            # 수식 계산 실패 시 data_only=True 폴백 여부 결정
            use_data_only = len(self._calculated_values) == 0
            if use_data_only:
                self._logger.info("수식 계산 결과 없음 → data_only=True 폴백 사용")

            # 4단계: 이미지 추출 (read_only=False로 열어야 이미지 접근 가능)
            update_progress(4, "이미지 추출 중...")
            self._extract_images(file_path)

            # 5단계: openpyxl로 구조 로드
            # - 수식 계산 성공: data_only=False (수식 텍스트 유지, 계산값은 캐시에서)
            # - 수식 계산 실패: data_only=True (엑셀 파일에 캐시된 값 사용)
            update_progress(5, "데이터 로드 중...")
            self._workbook = openpyxl.load_workbook(
                file_path, read_only=True, data_only=use_data_only
            )
            self._use_data_only = use_data_only
        except InvalidFileException as e:
            raise ExcelLoaderError(f"잘못된 엑셀 파일 형식입니다: {e}")
        except Exception as e:
            self._logger.error(f"파일 로드 실패: {e}")
            raise ExcelLoaderError(f"파일 로드 실패: {e}")

        self._file_path = file_path
        self._load_sheet()

    def _process_single_image(self, img_data: bytes, row: int, col: int) -> Tuple[str, Path, Path]:
        """단일 이미지 처리 (병렬 처리용)

        Args:
            img_data: 이미지 바이트 데이터
            row: 행 번호 (1-based)
            col: 열 번호

        Returns:
            (cell_key, 이미지 경로, 썸네일 경로)
        """
        cell_key = f"{row}_{col}"
        img_path = self._images_dir / f"img_{row}_{col}.png"
        thumb_path = self._images_dir / f"thumb_{row}_{col}.png"

        # PIL로 이미지 처리 및 저장
        pil_img = Image.open(BytesIO(img_data))
        pil_img.save(img_path, "PNG")

        # 썸네일 생성 및 저장
        thumb_img = pil_img.copy()
        thumb_img.thumbnail((self.THUMBNAIL_SIZE, self.THUMBNAIL_SIZE), Image.Resampling.LANCZOS)
        thumb_img.save(thumb_path, "PNG")

        return cell_key, img_path, thumb_path

    def _extract_images(self, file_path: Path) -> None:
        """엑셀 파일에서 이미지 추출 (병렬 처리 + 썸네일 미리 생성)"""
        t0 = time.time()
        self._image_map = {}
        self._thumbnail_map = {}

        try:
            # read_only=False로 열어야 이미지 접근 가능
            t_open = time.time()
            wb = openpyxl.load_workbook(file_path, read_only=False)
            self._logger.info(f"이미지용 워크북 열기: {time.time() - t_open:.2f}초")

            # 대상 시트 선택
            if self.DEFAULT_SHEET_NAME in wb.sheetnames:
                sheet = wb[self.DEFAULT_SHEET_NAME]
            else:
                sheet = wb.active

            images = sheet._images
            if not images:
                self._logger.info("이미지 없음")
                wb.close()
                return

            # 이미지 디렉토리 생성
            self._images_dir.mkdir(parents=True, exist_ok=True)

            # 이미지 데이터 수집 (메인 스레드에서)
            t_collect = time.time()
            image_tasks = []
            for img in images:
                try:
                    anchor = img.anchor
                    if hasattr(anchor, '_from'):
                        col = anchor._from.col
                        row = anchor._from.row + 1  # 0-based to 1-based
                        img_data = img._data()
                        image_tasks.append((img_data, row, col))
                except Exception as e:
                    self._logger.warning(f"이미지 데이터 수집 실패: {e}")
            wb.close()
            self._logger.info(f"이미지 데이터 수집: {time.time() - t_collect:.2f}초, {len(image_tasks)}개")

            # 병렬 이미지 처리
            t_process = time.time()
            with ThreadPoolExecutor(max_workers=4) as executor:
                futures = {
                    executor.submit(self._process_single_image, img_data, row, col): (row, col)
                    for img_data, row, col in image_tasks
                }

                for future in as_completed(futures):
                    row, col = futures[future]
                    try:
                        cell_key, img_path, thumb_path = future.result()
                        self._image_map[cell_key] = img_path
                        self._thumbnail_map[cell_key] = thumb_path
                    except Exception as e:
                        self._logger.warning(f"이미지 처리 실패 ({row}, {col}): {e}")

            self._logger.info(f"이미지 처리(병렬): {time.time() - t_process:.2f}초")
            self._logger.info(f"이미지 추출 완료: {len(self._image_map)}개, 총 {time.time() - t0:.2f}초")

        except Exception as e:
            self._logger.warning(f"이미지 추출 중 오류: {e}")

    def _calculate_formulas(self, file_path: Path, update_progress: callable = None) -> None:
        """openpyxl 기반 내장 수식 평가기로 수식 계산

        Named Range 테이블(RULA_A 등)과 셀 참조를 해석하여
        INDEX, IF, MIN, MAX 등의 수식을 Python eval()로 계산합니다.
        외부 라이브러리 없이 동작하므로 PyInstaller 배포와 호환됩니다.
        """

        def progress(step: int, message: str):
            if update_progress:
                update_progress(step, message)

        try:
            t0 = time.time()
            progress(1, "엑셀 파일 읽기 중...")
            self._logger.info("엑셀 파일 읽기 중...")
            wb = openpyxl.load_workbook(file_path, read_only=False, data_only=False)

            # 대상 시트 선택
            if self.DEFAULT_SHEET_NAME in wb.sheetnames:
                ws = wb[self.DEFAULT_SHEET_NAME]
                sheet_name = self.DEFAULT_SHEET_NAME
            else:
                ws = wb.active
                sheet_name = ws.title

            self._logger.info(f"엑셀 파일 읽기 완료 ({time.time() - t0:.1f}초)")

            # Named Range 테이블 로드
            t1 = time.time()
            progress(2, "수식 계산 중...")
            tables: Dict[str, List[List[Any]]] = {}
            for name, defn in wb.defined_names.items():
                try:
                    parts = defn.attr_text.split("!")
                    ref_sheet = parts[0].strip("'")
                    if ref_sheet in wb.sheetnames:
                        table_ws = wb[ref_sheet]
                        data = []
                        for row in table_ws.iter_rows(values_only=True):
                            data.append(list(row))
                        tables[name] = data
                except Exception as e:
                    self._logger.debug(f"Named Range 로드 실패 ({name}): {e}")

            self._logger.info(f"Named Range {len(tables)}개 로드: {list(tables.keys())}")

            # 셀 값 수집 및 수식 목록 생성
            cell_values: Dict[str, Any] = {}
            formulas: Dict[str, str] = {}

            for row in ws.iter_rows():
                for cell in row:
                    col_letter = get_column_letter(cell.column)
                    cell_ref = f"{col_letter}{cell.row}"
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        formulas[cell_ref] = cell.value
                    elif cell.value is not None:
                        cell_values[cell_ref] = cell.value

            self._logger.info(f"비수식 셀 {len(cell_values)}개, 수식 셀 {len(formulas)}개")
            wb.close()

            # 수식 반복 계산 (의존성 순서 자동 해결)
            eval_namespace: Dict[str, Any] = {
                "__builtins__": {},
                "IF": _excel_if,
                "INDEX": _excel_index,
                "MIN": min,
                "MAX": max,
            }
            eval_namespace.update(tables)

            max_iterations = 10
            for iteration in range(max_iterations):
                resolved = 0
                for ref, formula in formulas.items():
                    if ref not in cell_values:
                        result = self._eval_formula(formula, cell_values, eval_namespace)
                        if result is not None:
                            cell_values[ref] = result
                            resolved += 1
                if resolved == 0:
                    break
                self._logger.debug(f"반복 {iteration + 1}: {resolved}개 수식 계산 완료")

            self._logger.info(f"수식 계산 완료 ({time.time() - t1:.1f}초)")

            # 미해결 수식 확인
            unresolved = [ref for ref in formulas if ref not in cell_values]
            if unresolved:
                self._logger.warning(f"미해결 수식 {len(unresolved)}개: {unresolved[:10]}")

            # SHEET_NAME!CELL 형식으로 변환
            t2 = time.time()
            progress(3, "값 변환 중...")
            self._calculated_values = {}
            for ref, value in cell_values.items():
                full_ref = f"{sheet_name}!{ref}".upper()
                self._calculated_values[full_ref] = value

            self._logger.info(f"값 변환 완료 ({time.time() - t2:.1f}초), 총 {len(self._calculated_values)}개 셀")
            self._logger.info(f"전체 소요 시간: {time.time() - t0:.1f}초")
        except Exception as e:
            self._logger.warning(f"수식 계산 중 오류 (무시하고 진행): {e}")
            self._logger.warning(f"수식 계산 traceback:\n{traceback.format_exc()}")
            self._calculated_values = {}

    @staticmethod
    def _eval_formula(formula: str, cell_values: Dict[str, Any],
                      namespace: Dict[str, Any]) -> Any:
        """단일 수식 계산"""
        expr = formula[1:]  # 선행 = 제거

        # <> → != 변환
        expr = expr.replace("<>", "!=")
        # 등호를 이중등호로 변환 (비교 연산용, <=, >=, != 제외)
        expr = _EQ_RE.sub("==", expr)

        # 셀 참조를 값으로 치환
        def replace_ref(match: re.Match) -> str:
            ref = match.group(0)
            val = cell_values.get(ref)
            if val is None:
                return _UNRESOLVED
            if isinstance(val, (int, float)):
                return str(val)
            return repr(str(val))

        expr = _CELL_REF_RE.sub(replace_ref, expr)

        # 의존하는 셀이 아직 미해결이면 건너뜀
        if _UNRESOLVED in expr:
            return None

        # & → + (문자열 연결)
        expr = expr.replace("&", "+")

        try:
            result = eval(expr, namespace)  # noqa: S307
            if isinstance(result, float) and result == int(result):
                result = int(result)
            return result
        except Exception:
            return None

    def _load_sheet(self) -> None:
        """시트 데이터 로드"""
        # Capture Data 시트 또는 첫 번째 시트 선택
        if self.DEFAULT_SHEET_NAME in self._workbook.sheetnames:
            self._sheet = self._workbook[self.DEFAULT_SHEET_NAME]
            sheet_name = self.DEFAULT_SHEET_NAME
        else:
            self._sheet = self._workbook.active
            sheet_name = self._sheet.title

        # 헤더 읽기 (첫 번째 행)
        rows = list(self._sheet.iter_rows(values_only=False))
        if not rows:
            self._headers = []
            self._data = []
            return

        # 헤더 추출 (첫 번째 행)
        self._headers = [str(cell.value) if cell.value is not None else "" for cell in rows[0]]

        # 데이터 행 읽기
        self._data = []
        self._data_by_index = []
        for row_idx, row in enumerate(rows[1:], start=2):  # 엑셀은 1-based, 헤더가 1행
            row_dict = {}
            row_list = []
            for col_idx, cell in enumerate(row):
                # 셀 값 가져오기 (수식인 경우 계산된 값 사용)
                value = self._get_cell_value(cell, sheet_name, row_idx, col_idx)
                row_list.append(value)
                if col_idx < len(self._headers):
                    row_dict[self._headers[col_idx]] = value
            self._data.append(row_dict)
            self._data_by_index.append(row_list)

    def _get_cell_value(self, cell, sheet_name: str, row: int, col: int) -> Any:
        """셀 값 가져오기 (수식인 경우 계산된 값, 이미지인 경우 경로 사용)"""
        # 이미지가 있는 셀인지 확인
        cell_key = f"{row}_{col}"
        if cell_key in self._image_map:
            return self._image_map[cell_key]

        # data_only=True 모드: openpyxl이 이미 캐시된 값을 반환
        if self._use_data_only:
            value = cell.value
            # float가 정수인 경우 int로 변환 (5.0 → 5)
            if isinstance(value, float) and value.is_integer():
                value = int(value)
            return value

        # 수식 셀인지 확인
        if cell.data_type == 'f' or (isinstance(cell.value, str) and cell.value.startswith('=')):
            # 계산된 값 조회
            col_letter = get_column_letter(col + 1)  # 0-based to 1-based
            cell_ref = f"{sheet_name}!{col_letter}{row}".upper()

            if cell_ref in self._calculated_values:
                return self._calculated_values[cell_ref]
            else:
                self._logger.debug(f"계산값 없음: {cell_ref}")
                return ""  # 계산 실패 시 빈 문자열 반환 (수식 텍스트 노출 방지)

        return cell.value

    def _ensure_loaded(self) -> None:
        """데이터가 로드되었는지 확인"""
        if not self.is_loaded:
            raise ExcelLoaderError("먼저 load()를 호출하여 파일을 로드해야 합니다")

    def get_headers(self) -> list[str]:
        """헤더 목록 반환

        Returns:
            컬럼 이름 목록
        """
        self._ensure_loaded()
        return self._headers.copy()

    def get_row(self, index: int) -> dict[str, Any]:
        """특정 행 데이터 반환

        Args:
            index: 행 인덱스 (0부터 시작)

        Returns:
            {컬럼명: 값} 딕셔너리

        Raises:
            ExcelLoaderError: 잘못된 인덱스인 경우
        """
        self._ensure_loaded()

        if index < 0 or index >= len(self._data):
            raise ExcelLoaderError(f"잘못된 행 인덱스입니다: {index} (범위: 0-{len(self._data) - 1})")

        return self._data[index].copy()

    def get_rows(self, indices: list[int]) -> list[dict[str, Any]]:
        """다중 행 데이터 반환

        Args:
            indices: 행 인덱스 목록

        Returns:
            딕셔너리 목록
        """
        return [self.get_row(i) for i in indices]

    def get_all_rows(self) -> list[dict[str, Any]]:
        """전체 행 데이터 반환

        Returns:
            모든 행의 딕셔너리 목록
        """
        self._ensure_loaded()
        return [row.copy() for row in self._data]

    def get_row_by_index(self, row_index: int) -> list[Any]:
        """인덱스 기반 행 데이터 반환 (중복 헤더 지원)

        Args:
            row_index: 행 인덱스 (0부터 시작)

        Returns:
            값 리스트 (컬럼 인덱스로 접근)
        """
        self._ensure_loaded()

        if row_index < 0 or row_index >= len(self._data_by_index):
            raise ExcelLoaderError(f"잘못된 행 인덱스입니다: {row_index}")

        return self._data_by_index[row_index].copy()

    def get_headers_with_index(self) -> list[tuple[int, str]]:
        """인덱스와 함께 헤더 목록 반환

        Returns:
            [(인덱스, 헤더명), ...] 리스트
        """
        self._ensure_loaded()
        return [(i, h) for i, h in enumerate(self._headers)]

    def get_all_rows_by_index(self) -> list[list[Any]]:
        """전체 행 데이터 인덱스 기반 반환

        Returns:
            모든 행의 리스트
        """
        self._ensure_loaded()
        return [row.copy() for row in self._data_by_index]
